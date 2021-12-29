import openpyxl
import numpy as np 
import matplotlib.pyplot as plt 
  
def generate_dictionary(filename, conv_type, env_name):

	workbook = openpyxl.load_workbook(filename)
	worksheet = workbook.active

	max_row = worksheet.max_row
	max_column = worksheet.max_column

	# Iterate the rows and columns
	metrics = []
	exp_count = 0

	if conv_type == '3DConv':
		for i in range (0, 2):
			temp_dict = {'Naive': [], 'Const_mem': [], 'Shared_mem': [],
						 'Stream': [], 'Const_shared': [], 'Const_stream': [], 
						 'Final': []}
			for j in range (0, max_row + 1):
				if (j == 9) or (j == 19) or (j == 29):
					exp_count += 1
					metrics.append(temp_dict)
					temp_dict = {'Naive': [], 'Const_mem': [], 'Shared_mem': [],
						 'Stream': [], 'Const_shared': [], 'Const_stream': [], 
						 'Final': []}
					j += 1
					if j == 30:
						break

				temp = str(worksheet.cell(j + 1, i*5 + 1).value)
				if (('Exp' not in temp) and ('data size' not in temp)):
					temp_dict[temp].append(float((worksheet.cell(j+1, 2 + i*5).value)[:-2]))
					temp_dict[temp].append(float((worksheet.cell(j+1, 3 + i*5).value)[:-2]))
					temp_dict[temp].append(float((worksheet.cell(j+1, 4 + i*5).value)[:-1]))
					if ((worksheet.cell(j+1, 5).value) == 0.0) or ((worksheet.cell(j+1, 5).value) == 3.0): 
						if '1.530' in str(worksheet.cell(j+1, 5 + i*5).value) \
							or '1.529' in str(worksheet.cell(j+1, 5 + i*5).value):
							temp_dict[temp].append(float(1.530))
						else:
							temp_dict[temp].append(float(worksheet.cell(j+1, 5 + i*5).value))
					else:
						temp_dict[temp].append(float((worksheet.cell(j+1, 5 + i*5).value)[:-1]))

	elif conv_type == '2DConv':
		flag = 0
		for i in range (0, 2):
			temp_dict = {'Naive': [], 'Const_mem': [], 'Shared_mem': [],
						 'Stream': [], 'Const_shared': [], 'Const_stream': [], 
						 'Final': []}
			for j in range (0, max_row + 1):
				if (j == 9) or (j == 19) or (j == 29) or (j == 35):
					exp_count += 1
					metrics.append(temp_dict)
					if j == 35:
						flag = 1
						break
					temp_dict = {'Naive': [], 'Const_mem': [], 'Shared_mem': [],
						 'Stream': [], 'Const_shared': [], 'Const_stream': [], 
						 'Final': []}
					j += 1
				
				if (flag == 1) and j == 25:
					metrics.append(temp_dict)
					break 
				
				if i == 1 and j > 20:
					temp = str(worksheet.cell(j + 1, 1 + 5 * i).value)
				else:
					temp = str(worksheet.cell(j + 1, 1).value)

				if (('Exp' not in temp) and ('problem size' not in temp)):
					temp_dict[temp].append(float((worksheet.cell(j+1, 2 + i*5).value)[:-2]))
					temp_dict[temp].append(float((worksheet.cell(j+1, 3 + i*5).value)[:-2]))
					temp_dict[temp].append(float((worksheet.cell(j+1, 4 + i*5).value)[:-1]))
					if worksheet.cell(j+1, 5).value == 0.0: 
						temp_dict[temp].append(float(worksheet.cell(j+1, 5 + i*5).value))
					else:
						temp_dict[temp].append(float((worksheet.cell(j+1, 5 + i*5).value)[:-1]))
				
	return metrics, exp_count

def plots_gpu_kernel_time(metrics, conv_type, environment, indx):
	experiments = ['Naive', 'Const_mem', 'Shared_mem', 'Stream', 'Const_shared',
		'Const_stream', 'Final']

	experiments_stream = ['Stream', 'Const_stream', 'Final']

	gpu_ex_times = []
	gpu_ex_plus_memcpy_times = []
	error = []
	if conv_type == '3DConv':
		for i in range (0, len(metrics[indx])):
			gpu_ex_times.append(metrics[indx][experiments[i]][0])
			gpu_ex_plus_memcpy_times.append(metrics[indx][experiments[i]][1])
			error.append(metrics[indx][experiments[i]][2])

	elif conv_type == '2DConv':
		if indx != 3 and indx != 6:
			for i in range (0, len(metrics[indx])):
				gpu_ex_times.append(metrics[indx][experiments[i]][0])
				gpu_ex_plus_memcpy_times.append(metrics[indx][experiments[i]][1])
				error.append(metrics[indx][experiments[i]][2])
		else:
			for i in range (0, len(experiments_stream)):
				gpu_ex_times.append(metrics[indx][experiments_stream[i]][0])
				gpu_ex_plus_memcpy_times.append(metrics[indx][experiments_stream[i]][1])
				error.append(metrics[indx][experiments_stream[i]][2])

	if (conv_type == '2DConv') and ((indx == 3) or (indx == 6)): 
		X_axis = np.arange(len(experiments_stream))
		plt.bar(X_axis - 0.2, gpu_ex_times, 0.4,  label = 'GPU Kernel Execution (ms)')
		plt.bar(X_axis + 0.2, gpu_ex_plus_memcpy_times, 0.4, label = 'GPU Kernel Launch + Execution + Memcpy ops (ms)')
		
		plt.xticks(X_axis, experiments_stream)
		plt.xlabel("Kernel Configurations")
		plt.ylabel("Total Elapsed Time in ms")
		plt.title("Execution Time for Different Kernel Configurations for " 
					+ str (conv_type) + "olution in " + str(environment))
		plt.legend()
		plt.show()

	else:
		X_axis = np.arange(len(experiments))
		plt.bar(X_axis - 0.2, gpu_ex_times, 0.4,  label = 'GPU Kernel Execution (ms)')
		plt.bar(X_axis + 0.2, gpu_ex_plus_memcpy_times, 0.4, label = 'GPU Kernel Launch + Execution + Memcpy ops (ms)')
		
		plt.xticks(X_axis, experiments)
		plt.xlabel("Kernel Configurations")
		plt.ylabel("Total Elapsed Time in ms")
		plt.title("Execution Time for Different Kernel Configurations for " 
					+ str (conv_type) + "olution in " + str(environment))
		plt.legend()
		plt.show()


colab2DConvMetrics = "metrics_colab(TitanK80)2D.xlsx"
geeforceGtx2DConvMetrics = "metrics_GeeForceGtx1650Desktop2D.xlsx"
colab3DConvMetrics = "metrics_colab(TitanK80)3D.xlsx"
geeforceGtx3DConvMetrics = "metrics_GeeForceGtx1650Desktop3D.xlsx"

'''
colab3DMetrics, _ = generate_dictionary(colab3DConvMetrics, '3DConv', 'colab')
for i in range(0, len(colab3DMetrics)):
	plots_gpu_kernel_time(colab3DMetrics, '3DConv', 'COLAB', i)
'''

'''
geeforce3DMetrics, _ = generate_dictionary(geeforceGtx3DConvMetrics, '3DConv', 'geeforce')
for i in range(0, len(geeforce3DMetrics)):
	plots_gpu_kernel_time(geeforce3DMetrics, '3DConv', 'GeeforceGTX 1650', i)
'''

'''
colab2DMetrics, _ = generate_dictionary(colab2DConvMetrics, '2DConv', 'colab')
for i in range(0, len(colab2DMetrics)):
	plots_gpu_kernel_time(colab2DMetrics, '2DConv', 'COLAB', i)
'''

'''
geeforce2DMetrics, _ = generate_dictionary(geeforceGtx2DConvMetrics, '2DConv', 'geeforce')
for i in range(0, len(geeforce2DMetrics)):
	plots_gpu_kernel_time(geeforce2DMetrics, '2DConv', 'GeeforceGTX 1650', i)
'''
